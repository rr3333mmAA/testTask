from openpyxl import Workbook, load_workbook


def append_data_to_excel(data):
    try:
        # Load existing workbook if it exists, otherwise create a new one
        wb = load_workbook("orders.xlsx")
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # If the worksheet is empty, add header row
    if ws.max_row == 1:
        ws.append(['Username', 'Address', 'Product(quantity)', 'Total'])

    # Add the data to the worksheet
    ws.append(data)

    # Save the workbook
    wb.save("orders.xlsx")
